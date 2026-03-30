#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
File: cis_benchmark_converter.py
Author: Nicole Kemp
Created: 2026-03-11
Last Updated: 2026-03-25
Documentation Created: 2026-03-11

Description
-----------
This script extracts security recommendations from CIS Benchmark PDF documents
and converts them into structured data formats such as CSV, Excel, or JSON.

CIS Benchmarks are typically distributed as PDF documents containing security
configuration guidance for operating systems, cloud platforms, applications,
and infrastructure components. Because these documents are not structured as
machine-readable data, manually reviewing and mapping controls for compliance
or auditing purposes can be time-consuming.

This script automates the process by:

- Parsing CIS benchmark PDFs using the pdfplumber library
- Identifying benchmark recommendation titles
- Extracting associated sections such as Description, Rationale, Audit,
  Remediation, and Impact
- Exporting the extracted controls into structured formats suitable for
  compliance reviews, hardening baselines, and audit evidence preparation

Features
--------
- Extracts benchmark recommendations automatically from CIS PDF documents
- Supports CSV, Excel, and JSON output formats
- Excel output includes formatting and compliance tracking features
- Allows configurable extraction start page to skip front matter
- Removes page numbering artifacts during parsing
- Displays extraction progress with a progress bar
- Provides configurable logging verbosity

Typical Use Cases
-----------------
- Converting CIS benchmark controls into a review spreadsheet
- Tracking implementation status of CIS security recommendations
- Mapping CIS controls to internal security standards
- Preparing security hardening checklists
- Supporting vulnerability management or configuration audits

Dependencies
------------
Required Python libraries:

- pdfplumber
- openpyxl
- tqdm

Install dependencies using:

    pip install pdfplumber openpyxl tqdm

Command Example
---------------
    python cis_benchmark_converter.py \
        -i ./CIS_Azure_Benchmark.pdf \
        -f excel \
        --start_page 22 \
        --log_level INFO

Output
------
Depending on the selected format, the script generates one of the following:

CSV:
    Pipe-delimited table suitable for importing into spreadsheets or data tools.

Excel:
    Formatted spreadsheet with:
    - compliance status dropdown
    - conditional formatting
    - table styling

JSON:
    Structured data suitable for integration with automation or compliance tooling.

Notes
-----
This tool extracts information from CIS Benchmark documents. Users must ensure
they respect CIS licensing terms when using or redistributing extracted content.
"""

import argparse
import csv
import json
import logging
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from tqdm import tqdm

# -----------------------------------------------------------------------------
# Global Constants and Regular Expressions
# -----------------------------------------------------------------------------

# Matches recommendation titles such as:
# "1.1.1 (L1) Title of Recommendation"
# "2.3.4 Title of Recommendation"
TITLE_PATTERN: re.Pattern = re.compile(r'^(\d+\.\d+(?:\.\d+)*)\s*(\(L\d+\))?\s*(.*)')

# Matches page number strings such as "Page 123"
PAGE_NUMBER_PATTERN: re.Pattern = re.compile(r'\bPage\s+\d+\b', re.IGNORECASE)

# Matches version strings on the first page of a CIS PDF. Handles formats:
#   v1.2.0 - 2024       (canonical CIS format)
#   v1 (2024)           (abbreviated)
#   Version 1.0         (written out)
#   1.0.0               (bare numeric, whole line)
VERSION_PATTERN: re.Pattern = re.compile(
    r'^(?:v[\d.\-]+\s*-\s*.+|v\d[\d.]*\s*\(.*\)|[Vv]ersion\s+[\d.]+|\d+\.\d+[\d.]*\s*$)',
    re.IGNORECASE,
)

# Standard sections commonly present in CIS recommendations
SECTIONS_WITHOUT_CIS: List[str] = [
    "Profile Applicability:",
    "Description:",
    "Rationale:",
    "Impact:",
    "Audit:",
    "Remediation:",
    "Default Value:",
    "References:",
    "Additional Information:",
]

# Default compliance status for new records
DEFAULT_COMPLIANCE_STATUS = "To Review"

# Lines to scan beyond a candidate title when confirming a CIS recommendation.
# CIS PDFs consistently place "Profile Applicability:" within a few lines of
# the title; 10 provides enough headroom without accepting false positives.
MAX_PROFILE_LOOKAHEAD: int = 10

# Pre-lowercased section headers for case-insensitive matching.
# Kept in sync with SECTIONS_WITHOUT_CIS; used alongside _normalize_section_line().
SECTIONS_LOWER: List[str] = [s.lower() for s in SECTIONS_WITHOUT_CIS]

# -----------------------------------------------------------------------------
# Utility Functions
# -----------------------------------------------------------------------------

def setup_logging(log_level: str) -> None:
    """
    Configure application logging.

    Args:
        log_level: Logging level string such as DEBUG, INFO, WARNING, ERROR.
    """
    numeric_level = getattr(logging, log_level.upper(), logging.INFO)
    logging.basicConfig(
        level=numeric_level,
        format="%(asctime)s | %(levelname)s | %(message)s"
    )


def remove_page_numbers(text: str) -> str:
    """
    Remove page number references like "Page 123" from a text string.

    Args:
        text: Input text.

    Returns:
        Cleaned text with page references removed.
    """
    return PAGE_NUMBER_PATTERN.sub("", text).strip()


def generate_unique_filename(base_name: str, extension: str) -> str:
    """
    Generate a unique output filename by appending a numeric suffix if needed.

    Args:
        base_name: File name without extension.
        extension: Output file extension without leading dot.

    Returns:
        A unique file path string.
    """
    file_path = Path(f"{base_name}.{extension}")
    counter = 1

    while file_path.exists():
        logging.debug(
            "Output file '%s' already exists; trying suffix (%s).", file_path, counter
        )
        file_path = Path(f"{base_name}({counter}).{extension}")
        counter += 1

    logging.debug("Resolved unique output filename: %s", file_path)
    return str(file_path)


def normalize_whitespace(text: str) -> str:
    """
    Collapse repeated whitespace into single spaces and trim outer whitespace.

    Args:
        text: Input text.

    Returns:
        Normalized text.
    """
    return re.sub(r"\s+", " ", text).strip()


def _print_extraction_stats(recommendations: List[Dict[str, str]]) -> None:
    """
    Print a per-section coverage report to stdout.

    For each standard CIS section, shows the percentage of recommendations
    where that section was successfully extracted with non-empty content.
    Helps users quickly assess extraction quality without opening the output file.

    Args:
        recommendations: Extracted recommendation list.
    """
    total = len(recommendations)
    if total == 0:
        return

    print(f"\nExtraction coverage ({total} recommendations):")
    for sec in SECTIONS_WITHOUT_CIS:
        key = sec[:-1]  # strip trailing colon to match stored dict key
        filled = sum(1 for r in recommendations if r.get(key, "").strip())
        pct = (filled / total) * 100
        print(f"  {key:<26} {pct:5.1f}%")


def _normalize_section_line(line: str) -> str:
    """
    Normalise a PDF line for section header matching.

    Lowercases the line and collapses any whitespace surrounding colons so
    that variants like "Description :", "DESCRIPTION:" and "description:"
    all compare equal to the entries in SECTIONS_LOWER.

    Args:
        line: Raw text line from the PDF.

    Returns:
        Normalised line string.
    """
    return re.sub(r'\s*:\s*', ':', line.lower())


# -----------------------------------------------------------------------------
# PDF Extraction Functions
# -----------------------------------------------------------------------------

def extract_title_and_version(input_file: Path) -> Tuple[str, str]:
    """
    Extract the document title and version from the first page of the PDF.

    Logic:
    - Reads the first page
    - Treats lines before a version-looking line as title content
    - Attempts to identify a version line such as "v1.2.0 - 2024"

    Args:
        input_file: Path to the input PDF.

    Returns:
        Tuple of (title, version). Version may be an empty string if not found.
    """
    logging.debug("Opening PDF for title/version extraction: %s", input_file)

    try:
        with pdfplumber.open(str(input_file)) as pdf:
            if pdf.is_encrypted:
                raise ValueError(
                    f"PDF '{input_file}' is encrypted. Decrypt it before processing."
                )
            first_page_text = pdf.pages[0].extract_text()
            if not first_page_text:
                logging.warning("First page contained no extractable text.")
                return "CIS Benchmark Document", ""
            lines = first_page_text.splitlines()
            logging.debug("First page yielded %d lines for title extraction.", len(lines))
    except Exception as exc:
        logging.error("Error opening PDF for title extraction: %s", exc)
        raise

    title_lines: List[str] = []
    version = ""

    for line in lines:
        clean_line = line.strip()
        if VERSION_PATTERN.match(clean_line):
            logging.debug("Version line detected: '%s'", clean_line)
            version = clean_line
            break
        if clean_line:
            logging.debug("Title candidate line: '%s'", clean_line)
            title_lines.append(clean_line)

    title = " ".join(title_lines) if title_lines else "CIS Benchmark Document"

    if not title_lines:
        logging.warning("No title lines found on first page; using default title.")

    logging.debug("Extracted title: '%s' | version: '%s'", title, version)
    return normalize_whitespace(title), version


def read_pdf(input_file: Path, start_page: int = 10) -> str:
    """
    Read text from a PDF file starting at a specified page.

    Args:
        input_file: Path to the input PDF.
        start_page: 1-based page number to begin extraction.

    Raises:
        ValueError: If start_page is invalid.
        Exception: If PDF reading fails.

    Returns:
        A single concatenated string of extracted text.
    """
    logging.info("Reading PDF from page %s onwards...", start_page)

    if start_page < 1:
        raise ValueError("start_page must be greater than or equal to 1.")

    try:
        with pdfplumber.open(str(input_file)) as pdf:
            if pdf.is_encrypted:
                raise ValueError(
                    f"PDF '{input_file}' is encrypted. Decrypt it before processing."
                )
            total_pages = len(pdf.pages)
            logging.debug("PDF opened: %d total pages detected.", total_pages)

            if start_page > total_pages:
                logging.error(
                    "start_page (%d) exceeds total page count (%d).",
                    start_page, total_pages
                )
                raise ValueError(
                    f"Start page {start_page} exceeds total page count ({total_pages})."
                )

            pages_to_read = total_pages - start_page + 1
            logging.debug("Extracting %d pages (pages %d to %d).", pages_to_read, start_page, total_pages)

            text_pages: List[str] = []
            empty_pages: List[int] = []

            for page in tqdm(
                pdf.pages[start_page - 1:],
                desc="Extracting pages",
                unit="page",
                total=pages_to_read,
            ):
                page_number = page.page_number
                page_text = page.extract_text() or ""
                if page_text:
                    text_pages.append(page_text)
                else:
                    empty_pages.append(page_number)
                    logging.debug("Page %d yielded no extractable text.", page_number)

    except ValueError:
        raise
    except Exception as exc:
        logging.error("Failed to read PDF '%s': %s", input_file, exc)
        raise

    if empty_pages:
        logging.warning(
            "%d page(s) yielded no text and were skipped: %s",
            len(empty_pages),
            empty_pages,
        )

    extracted_chars = sum(len(p) for p in text_pages)
    logging.debug(
        "PDF extraction complete: %d pages with content, %d characters total.",
        len(text_pages), extracted_chars
    )

    return "\n".join(text_pages)


def find_profile_applicability(
    lines: List[str],
    start_index: int,
    max_depth: int = MAX_PROFILE_LOOKAHEAD
) -> bool:
    """
    Determine whether a detected title line is likely a valid recommendation.

    The script uses the presence of "Profile Applicability:" shortly after the
    title as an indicator that the title belongs to a real CIS recommendation.

    Args:
        lines: All text lines from the PDF.
        start_index: Index of the candidate title line.
        max_depth: Number of subsequent lines to inspect.

    Returns:
        True if "Profile Applicability:" is found before another section or title.
    """
    end_index = min(start_index + max_depth, len(lines))
    candidate_line = lines[start_index].strip()
    logging.debug(
        "Checking profile applicability for candidate at line %d: '%s'",
        start_index, candidate_line
    )

    for i in range(start_index + 1, end_index):
        line = remove_page_numbers(lines[i].strip())
        norm_line = _normalize_section_line(line)

        if norm_line.startswith("profile applicability:"):
            logging.debug("Profile Applicability found at line %d - candidate accepted.", i)
            return True

        if TITLE_PATTERN.match(line) or any(norm_line.startswith(sec) for sec in SECTIONS_LOWER):
            logging.debug(
                "Boundary reached at line %d ('%s') before Profile Applicability — candidate rejected.",
                i, line
            )
            return False

    logging.debug(
        "Profile Applicability not found within %d lines of candidate — candidate rejected.",
        max_depth
    )
    return False


def extract_section(lines: List[str], start_index: int, section_name: str) -> Tuple[str, int]:
    """
    Extract the content of a section until a boundary is reached.

    Boundaries include:
    - another known section header
    - a new recommendation title
    - a line starting with "CIS Controls"

    Args:
        lines: All text lines from the PDF.
        start_index: Index where the section header was found.
        section_name: Name of the section being extracted.

    Returns:
        Tuple of (section_content, next_index)
    """

    logging.debug(
        "Extracting section '%s' starting at line %d.", section_name, start_index
    )

    content: List[str] = []
    current_index = start_index + 1

    while current_index < len(lines):
        line = remove_page_numbers(lines[current_index].strip())
        norm_line = _normalize_section_line(line)

        if (
            any(norm_line.startswith(sec) for sec in SECTIONS_LOWER)
            or TITLE_PATTERN.match(line)
            or norm_line.startswith("cis controls")
        ):
            logging.debug(
                "Section '%s' boundary at line %d: '%s'.",
                section_name, current_index, line
            )
            break

        if line:
            content.append(line)

        current_index += 1

    result = normalize_whitespace(" ".join(content))
    logging.debug(
        "Section '%s' extracted %d content lines (%d chars).",
        section_name, len(content), len(result)
    )
    return result, current_index


def extract_recommendations(full_text: str) -> List[Dict[str, str]]:
    """
    Parse extracted PDF text into structured CIS recommendations.

    Args:
        full_text: Concatenated text extracted from the PDF.

    Returns:
        List of recommendation dictionaries.
    """
    recommendations: List[Dict[str, str]] = []
    lines = full_text.splitlines()
    current_recommendation: Dict[str, str] = {}
    current_index = 0

    logging.debug("Beginning recommendation extraction over %d lines.", len(lines))

    while current_index < len(lines):
        line = remove_page_numbers(lines[current_index].strip())

        title_match = TITLE_PATTERN.match(line)
        if title_match:
            if find_profile_applicability(lines, current_index):
                if current_recommendation:
                    logging.debug(
                        "Finalising recommendation %s before starting new one.",
                        current_recommendation.get("Number", "?")
                    )
                    recommendations.append(current_recommendation)

                current_recommendation = {
                    "Number": title_match.group(1),
                    "Level": title_match.group(2) or "",
                    "Title": title_match.group(3).strip(),
                }
                logging.debug(
                    "New recommendation detected — Number: %s, Level: %s, Title start: '%s'",
                    current_recommendation["Number"],
                    current_recommendation["Level"] or "(none)",
                    current_recommendation["Title"][:60],
                )

                # Capture multi-line titles until another section header or title appears
                while current_index + 1 < len(lines):
                    next_line = remove_page_numbers(lines[current_index + 1].strip())
                    norm_next = _normalize_section_line(next_line)

                    if (
                        any(norm_next.startswith(sec) for sec in SECTIONS_LOWER)
                        or TITLE_PATTERN.match(next_line)
                    ):
                        break

                    if next_line:
                        logging.debug("Multi-line title continuation: '%s'", next_line[:60])
                        current_recommendation["Title"] += f" {next_line}"

                    current_index += 1

                current_recommendation["Title"] = normalize_whitespace(
                    current_recommendation["Title"]
                )
                logging.debug(
                    "Finalised title for %s: '%s'",
                    current_recommendation["Number"],
                    current_recommendation["Title"][:80],
                )
            else:
                logging.debug(
                    "Line %d matched title pattern but was rejected (no Profile Applicability): '%s'",
                    current_index, line[:80]
                )

        norm_line = _normalize_section_line(line)
        for sec, sec_lower in zip(SECTIONS_WITHOUT_CIS, SECTIONS_LOWER):
            if norm_line.startswith(sec_lower) and current_recommendation:
                logging.debug(
                    "Section header '%s' found at line %d for recommendation %s.",
                    sec, current_index, current_recommendation.get("Number", "?")
                )
                content, next_index = extract_section(lines, current_index, sec)
                current_recommendation[sec[:-1]] = content
                current_index = next_index - 1
                break

        current_index += 1

    if current_recommendation:
        logging.debug(
            "Appending final recommendation %s.", current_recommendation.get("Number", "?")
        )
        recommendations.append(current_recommendation)

    logging.debug("Raw recommendation count before deduplication: %d.", len(recommendations))

    # Deduplicate by (Number, Title) while preserving most recent occurrence
    unique_recommendations = {
        (rec.get("Number", ""), rec.get("Title", "")): rec
        for rec in recommendations
    }

    duplicates_removed = len(recommendations) - len(unique_recommendations)
    if duplicates_removed:
        logging.warning(
            "%d duplicate recommendation(s) removed during deduplication.", duplicates_removed
        )
    else:
        logging.debug("No duplicates found.")

    final_recommendations = list(unique_recommendations.values())
    logging.info("Extracted %d unique recommendations.", len(final_recommendations))
    return final_recommendations


# -----------------------------------------------------------------------------
# Output Generation
# -----------------------------------------------------------------------------

def build_headers() -> List[str]:
    """
    Build the ordered output headers for CSV and Excel exports.

    Returns:
        List of column headers.
    """
    headers = ["Compliance Status", "Number", "Level", "Title"]
    headers.extend(sec[:-1] for sec in SECTIONS_WITHOUT_CIS)
    return headers


def _build_output_row(recommendation: Dict[str, str], headers: List[str]) -> List[str]:
    """
    Build a single output row from a recommendation dict.

    Args:
        recommendation: Recommendation data dict.
        headers: Ordered list of column headers.

    Returns:
        List of cell values in header order.
    """
    return [
        recommendation.get(h, "") or (DEFAULT_COMPLIANCE_STATUS if h == "Compliance Status" else "")
        for h in headers
    ]


def write_csv(
    recommendations: List[Dict[str, str]],
    output_file: Path,
    headers: List[str],
    title: str,
    version: str
) -> None:
    """
    Write recommendations to a pipe-delimited CSV file.
    """
    logging.debug("Opening CSV output file for writing: %s", output_file)

    try:
        with output_file.open(mode="w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file, delimiter="|")
            writer.writerow([title if title else "CIS Benchmark Document"])
            writer.writerow([version if version else ""])
            writer.writerow([])
            writer.writerow(headers)
            logging.debug("CSV header row written: %d columns.", len(headers))

            for idx, recommendation in enumerate(recommendations):
                writer.writerow(_build_output_row(recommendation, headers))
                logging.debug("CSV row %d written for recommendation %s.", idx + 1, recommendation.get("Number", "?"))

    except OSError as exc:
        logging.error("Failed to write CSV file '%s': %s", output_file, exc)
        raise

    logging.debug("CSV write complete: %d data rows written.", len(recommendations))


def apply_excel_formatting(
    sheet,
    headers: List[str],
    recommendation_count: int
) -> None:
    """
    Apply data validation, conditional formatting, table styling, and widths
    to the Excel worksheet.

    Args:
        sheet: Openpyxl worksheet object.
        headers: Header list.
        recommendation_count: Number of recommendation rows.
    """
    start_row = 5
    end_row = recommendation_count + start_row - 1
    logging.debug(
        "Applying Excel formatting: %d recommendations, data rows %d to %d.",
        recommendation_count, start_row, end_row
    )

    try:
        # Data validation for compliance status
        dv = DataValidation(
            type="list",
            formula1='"Compliant,Non-Compliant,To Review"',
            showDropDown=False
        )
        sheet.add_data_validation(dv)
        logging.debug("Data validation added for column A rows %d to %d.", start_row, end_row)

        for row_idx in range(start_row, end_row + 1):
            dv.add(sheet[f"A{row_idx}"])

        # Conditional formatting fills
        compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        non_compliant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        to_review_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Relative formulas applied over the range
        compliant_rule = FormulaRule(formula=['$A5="Compliant"'], fill=compliant_fill)
        non_compliant_rule = FormulaRule(formula=['$A5="Non-Compliant"'], fill=non_compliant_fill)
        to_review_rule = FormulaRule(formula=['$A5="To Review"'], fill=to_review_fill)

        if recommendation_count > 0:
            cf_range = f"A{start_row}:A{end_row}"
            sheet.conditional_formatting.add(cf_range, compliant_rule)
            sheet.conditional_formatting.add(cf_range, non_compliant_rule)
            sheet.conditional_formatting.add(cf_range, to_review_rule)
            logging.debug("Conditional formatting rules applied to range %s.", cf_range)
        else:
            logging.warning("No recommendations present; conditional formatting skipped.")

        # Create Excel table
        last_column = get_column_letter(len(headers))
        table_range = f"A4:{last_column}{max(4, end_row)}"
        logging.debug("Creating Excel table over range %s.", table_range)

        table = Table(displayName="CISRecommendations", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

        # Column widths
        width_map = {
            "A": 18,
            "B": 12,
            "C": 10,
            "D": 60,
        }

        for col_letter, width in width_map.items():
            sheet.column_dimensions[col_letter].width = width

        for col_idx in range(5, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 30

        logging.debug("Column widths set for %d columns.", len(headers))

    except Exception as exc:
        logging.error("Failed to apply Excel formatting: %s", exc)
        raise


def write_excel(
    recommendations: List[Dict[str, str]],
    output_file: Path,
    headers: List[str],
    title: str,
    version: str
) -> None:
    """
    Write recommendations to a formatted Excel workbook.
    """
    logging.debug("Building Excel workbook with %d recommendations.", len(recommendations))

    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Recommendations"

        # Title and version metadata
        sheet["A1"] = title if title else "CIS Benchmark Document"
        sheet["A1"].font = Font(size=14, bold=True)

        sheet["A2"] = version if version else ""
        sheet["A2"].font = Font(size=12, italic=True)
        logging.debug("Excel metadata written — title: '%s', version: '%s'.", title, version)

        # Spacer row
        sheet.append([""] * len(headers))

        # Header row at Excel row 4
        sheet.append(headers)
        logging.debug("Excel header row written at row 4: %d columns.", len(headers))

        for idx, recommendation in enumerate(recommendations):
            sheet.append(_build_output_row(recommendation, headers))
            logging.debug(
                "Excel row %d written for recommendation %s.",
                idx + 5, recommendation.get("Number", "?")
            )

        apply_excel_formatting(sheet, headers, len(recommendations))

        logging.debug("Saving Excel workbook to: %s", output_file)
        workbook.save(str(output_file))

    except OSError as exc:
        logging.error("Failed to save Excel file '%s': %s", output_file, exc)
        raise

    logging.debug("Excel write complete: %d data rows written.", len(recommendations))


def write_json(
    recommendations: List[Dict[str, str]],
    output_file: Path,
    title: str,
    version: str
) -> None:
    """
    Write recommendations to a JSON file.
    """
    logging.debug("Building JSON structure for %d recommendations.", len(recommendations))

    data = {
        "document_title": title if title else "CIS Benchmark Document",
        "document_version": version,
        "recommendations": recommendations,
    }

    try:
        with output_file.open("w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
    except OSError as exc:
        logging.error("Failed to write JSON file '%s': %s", output_file, exc)
        raise

    logging.debug("JSON write complete: %d recommendations serialised.", len(recommendations))


def write_output(
    recommendations: List[Dict[str, str]],
    output_file: Path,
    output_format: str,
    title: str,
    version: str
) -> None:
    """
    Write extracted recommendations to the requested output format.

    Args:
        recommendations: Extracted recommendation records.
        output_file: Target output path.
        output_format: csv, excel, or json.
        title: Extracted document title.
        version: Extracted document version.
    """
    logging.info("Writing output to %s in %s format...", output_file, output_format.upper())

    headers = build_headers()

    try:
        if output_format == "csv":
            write_csv(recommendations, output_file, headers, title, version)
        elif output_format == "excel":
            write_excel(recommendations, output_file, headers, title, version)
        elif output_format == "json":
            write_json(recommendations, output_file, title, version)
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
    except Exception as exc:
        logging.error("Failed to write output: %s", exc)
        raise

    logging.info("Finished writing %s recommendations to %s.", len(recommendations), output_file)


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def parse_arguments() -> argparse.Namespace:
    """
    Parse command-line arguments.

    Returns:
        Parsed argument namespace.
    """
    parser = argparse.ArgumentParser(
        description="Extract CIS Benchmark recommendations from a PDF into CSV, Excel, or JSON.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    parser.add_argument(
        "-i", "--input",
        required=True,
        metavar="PDF_PATH",
        help="Path to the CIS Benchmark PDF file.",
    )

    parser.add_argument(
        "-o", "--output",
        default=None,
        metavar="OUTPUT_PATH",
        help=(
            "Output file path without extension. "
            "Defaults to the input filename stem in the current directory."
        ),
    )

    parser.add_argument(
        "-f", "--format",
        dest="output_format",
        choices=["csv", "excel", "json"],
        default="excel",
        help="Output format.",
    )

    parser.add_argument(
        "--start_page",
        type=int,
        default=10,
        metavar="N",
        help="1-based page number to begin extraction from (skip front matter).",
    )

    parser.add_argument(
        "--log_level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        default="INFO",
        help="Logging verbosity level.",
    )

    return parser.parse_args()


def resolve_output_path(args: argparse.Namespace) -> Path:
    """
    Determine the output file path based on parsed arguments.

    If --output is not supplied, derives the base name from the input filename.

    Args:
        args: Parsed command-line arguments.

    Returns:
        Resolved output Path object (without extension).
    """
    if args.output:
        base_name = args.output
        logging.debug("Output base name provided explicitly: '%s'.", base_name)
    else:
        base_name = Path(args.input).stem
        logging.debug("Output base name derived from input stem: '%s'.", base_name)

    extension_map = {"csv": "csv", "excel": "xlsx", "json": "json"}
    extension = extension_map[args.output_format]

    resolved = Path(generate_unique_filename(base_name, extension))
    logging.debug("Resolved output path: %s", resolved)
    return resolved


def main() -> None:
    """
    Entry point for the CIS Benchmark Converter.

    Orchestrates argument parsing, PDF extraction, and output generation.
    """
    args = parse_arguments()
    setup_logging(args.log_level)

    logging.debug(
        "Arguments: input='%s', output='%s', format='%s', start_page=%d, log_level='%s'",
        args.input, args.output, args.output_format, args.start_page, args.log_level
    )

    input_file = Path(args.input)

    if not input_file.exists():
        logging.error("Input file not found: %s", input_file)
        raise SystemExit(1)

    if input_file.suffix.lower() != ".pdf":
        logging.error(
            "Input file does not appear to be a PDF (got '%s'): %s",
            input_file.suffix, input_file
        )
        raise SystemExit(1)

    logging.debug("Input file validated: %s (%d bytes).", input_file, input_file.stat().st_size)

    output_file = resolve_output_path(args)

    if output_file.parent != Path("."):
        logging.debug("Creating output directory: %s", output_file.parent)
        output_file.parent.mkdir(parents=True, exist_ok=True)

    logging.info("Input:       %s", input_file)
    logging.info("Output:      %s", output_file)
    logging.info("Format:      %s", args.output_format.upper())
    logging.info("Start page:  %s", args.start_page)

    title, version = extract_title_and_version(input_file)
    logging.info("Document title:   %s", title)
    logging.info("Document version: %s", version or "(not detected)")

    full_text = read_pdf(input_file, start_page=args.start_page)
    logging.debug("Total characters extracted from PDF: %d.", len(full_text))

    if not full_text.strip():
        logging.error(
            "No text could be extracted from '%s'. "
            "The file may be image-based (scanned PDF) or --start_page (%d) may be too high.",
            input_file, args.start_page
        )
        raise SystemExit(1)

    recommendations = extract_recommendations(full_text)

    if not recommendations:
        logging.warning(
            "No recommendations were extracted. "
            "Try adjusting --start_page or verify the PDF is a CIS Benchmark document."
        )
    else:
        _print_extraction_stats(recommendations)

    write_output(recommendations, output_file, args.output_format, title, version)

    print(f"\nExtraction complete: {len(recommendations)} recommendations written to {output_file}")


if __name__ == "__main__":
    main()
